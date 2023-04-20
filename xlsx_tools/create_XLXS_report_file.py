from openpyxl import load_workbook
import os
from must_have.make_documents_subfolder import make_documents_subfolder
from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font
)
from openpyxl import Workbook
import string


def create_report_file(report_folder, file_name, sheet_name, list_with_columns_names):
    report_file_name = f"{file_name}.xlsx"
    path_to_file = f'{report_folder}/{report_file_name}'

    if os.path.exists(path_to_file):
        wb = load_workbook(path_to_file)  # файл есть и открываю его
        ws = wb.create_sheet(sheet_name)  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файа еще нет
        ws.title = sheet_name  # если файда еще нет

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 10

    ws.row_dimensions[3].height = 40
    ws.row_dimensions[4].height = 40
    ws.row_dimensions[5].height = 40

    capital_letters = string.ascii_uppercase
    for x in range(len(list_with_columns_names)):
        ws.column_dimensions[f'{capital_letters[x]}'].width = 50  # задаю ширину колонки
        ws[f'{capital_letters[x]}1'] = list_with_columns_names[x]
        ws[f'{capital_letters[x]}1'].font = Font(size=18, bold=True)
        ws[f'{capital_letters[x]}1'].alignment = Alignment(horizontal='center')
        ws[f'{capital_letters[x]}1'].fill = PatternFill('solid', fgColor="DDDDDD")

        thins = Side(border_style="double", color="1d51e7")
        ws[f'{capital_letters[x]}1'].border = Border(top=thins, bottom=thins, left=thins, right=thins)

    wb.save(path_to_file)

    return path_to_file


def create_report(subfolder_name, file_name, sheet_name, list_with_columns_names):  # return path to file
    return create_report_file(make_documents_subfolder(subfolder_name), file_name, sheet_name, list_with_columns_names)


if __name__ == '__main__':
    # print(create_report("SMT_clinic", "appointments", 'clinic', ['doctor', 'appoinment_date', 'comments', 'ygjhghg']))
    print(create_report("Kommersant", "deleted_images", "shoot_id", ['shoot_id', "action"]))
