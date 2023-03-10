import openpyxl
from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font
)


def write_to_xlsx(all_data, common_days, path_to_file):
    common_days = sorted([x for x in common_days])  # преобразую коллекцию в отсортированный список
    wb = openpyxl.load_workbook(path_to_file, read_only=False)
    sheet = wb.active
    doctors = [x for x in all_data]
    for doctor in enumerate(doctors):
        for day in enumerate(common_days):
            sheet.cell(row=3 + doctor[0], column=1).alignment = Alignment(horizontal='center')
            sheet.cell(row=3 + doctor[0], column=1).font = Font(size=14, bold=True)
            sheet.cell(row=3 + doctor[0], column=1).value = doctor[1]
            sheet.cell(row=3 + doctor[0], column=2 + day[0]).alignment = Alignment(wrap_text=True)
            sheet.cell(row=3 + doctor[0], column=2 + day[0]).value = ", ".join(all_data[doctor[1]][day[1]])
    wb.save(path_to_file)


if __name__ == '__main__':
    write_to_xlsx({'Гаглоева': {'7': ['13:40', '14:00', '14:20'],
                                '9': ['10:00', '10:20', '10:40', '11:00', '11:20', '11:40', '12:00', '12:20', '12:40',
                                      '13:00', '13:20', '13:40', '14:00', '14:20'],
                                '11': ['10:00', '10:20', '11:40', '12:20', '12:40', '13:00', '13:20', '13:40', '14:00',
                                       '14:20', '14:40', '15:00', '15:20', '15:40'],
                                '14': ['08:40', '09:00', '09:20', '09:40', '10:00', '10:20', '10:40', '11:00', '11:20',
                                       '11:40', '12:00', '12:20', '12:40', '13:00', '13:20', '13:40', '14:00', '14:20',
                                       '14:40'],
                                '16': ['08:00', '08:20', '08:40', '09:00', '09:20', '09:40', '10:00', '10:20', '10:40',
                                       '11:20', '11:40', '12:00', '12:20', '12:40', '13:00', '13:40', '14:00', '14:20',
                                       '14:40'],
                                '18': ['09:00', '09:20', '09:40', '10:00', '10:20', '10:40', '11:00', '11:20', '11:40',
                                       '12:00', '12:20', '12:40', '13:00', '13:20', '13:40', '14:00', '14:20', '14:40',
                                       '15:00', '15:20', '15:40'],
                                '21': ['08:00', '08:20', '08:40', '09:00', '09:20', '09:40', '10:00', '10:20', '10:40',
                                       '11:00', '11:20', '11:40', '12:00', '12:20', '12:40', '13:00', '13:20', '13:40',
                                       '14:00'],
                                '23': ['08:00', '08:20', '08:40', '09:00', '09:20', '09:40', '10:00', '10:20', '10:40',
                                       '11:00', '11:20', '11:40', '12:00', '12:20', '12:40', '13:00', '13:20', '13:40',
                                       '14:00', '14:20', '14:40'],
                                '25': ['09:00', '09:20', '09:40', '10:00', '10:20', '11:00', '11:20', '11:40', '12:00',
                                       '12:20', '12:40', '13:00', '13:20', '13:40', '14:00', '14:20', '14:40', '15:00',
                                       '15:20', '15:40'],
                                '28': ['08:00', '08:20', '08:40', '09:00', '09:20', '09:40', '10:00', '10:20', '10:40',
                                       '11:00', '11:20', '11:40', '12:00', '12:20', '12:40', '13:00', '13:20', '13:40',
                                       '14:00', '14:20', '14:40'],
                                '30': ['08:00', '08:20', '08:40', '09:00', '09:20', '09:40', '10:00', '10:20', '10:40',
                                       '11:00', '11:20', '11:40', '12:00', '12:20', '12:40', '13:00', '13:20', '13:40',
                                       '14:00', '14:20', '14:40'],
                                '1': ['09:00', '09:20', '09:40', '10:00', '10:20', '10:40', '11:00', '11:20', '11:40',
                                      '12:00', '12:20', '12:40', '13:00', '13:20', '13:40', '14:00', '14:20', '14:40',
                                      '15:00', '15:20', '15:40']},
                   'Марченко': {'7': ['08:00', '09:00', '09:30', '10:30', '11:30', '13:00', '13:30'],
                                '10': ['15:00', '15:30', '16:00', '19:30', '20:00', '20:30'],
                                '14': ['08:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '13:00', '13:30'],
                                '17': ['15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00',
                                       '19:30', '20:00', '20:30'],
                                '21': ['08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00',
                                       '12:30', '13:00', '13:30'],
                                '24': ['15:00', '15:30', '16:00', '16:30', '18:00', '18:30', '19:00', '19:30', '20:00',
                                       '20:30'],
                                '28': ['08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00',
                                       '12:30', '13:00', '13:30'],
                                '31': ['15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00',
                                       '20:00', '20:30']}, 'Тарасова': {'9': ['09:00', '11:30', '12:00'],
                                                                        '13': ['11:30', '13:00', '13:30', '14:00',
                                                                               '14:30', '15:30', '16:30', '17:30'],
                                                                        '14': ['09:00', '09:30', '10:30', '11:00',
                                                                               '11:30', '12:00', '12:30', '13:00',
                                                                               '13:30', '14:00', '19:00'],
                                                                        '15': ['16:00', '16:30', '17:30', '18:00',
                                                                               '18:30', '19:00'],
                                                                        '16': ['08:30', '09:00', '09:30', '10:00',
                                                                               '10:30', '11:30', '12:00', '12:30',
                                                                               '13:00', '13:30', '14:00'],
                                                                        '20': ['14:30', '15:30', '16:00', '16:30',
                                                                               '17:00', '18:00', '18:30', '19:00',
                                                                               '20:00', '20:30'],
                                                                        '21': ['08:00', '08:30', '09:00', '09:30',
                                                                               '10:00', '11:00', '11:30', '12:30',
                                                                               '13:00', '13:30', '14:00', '14:30',
                                                                               '15:00', '15:30', '16:00', '16:30',
                                                                               '18:30', '19:30'],
                                                                        '22': ['14:30', '15:00', '15:30', '16:30',
                                                                               '17:00', '17:30', '18:00', '18:30',
                                                                               '19:00', '20:00', '20:30'],
                                                                        '23': ['08:00', '08:30', '09:00', '09:30',
                                                                               '10:00', '10:30', '11:00', '11:30',
                                                                               '12:00', '12:30', '13:00', '13:30',
                                                                               '14:00'],
                                                                        '27': ['14:30', '15:30', '16:00', '16:30',
                                                                               '17:00', '17:30', '18:00', '18:30',
                                                                               '19:30', '20:00', '20:30'],
                                                                        '28': ['08:00', '08:30', '09:00', '09:30',
                                                                               '10:00', '10:30', '11:00', '11:30',
                                                                               '12:00', '12:30', '13:00', '13:30',
                                                                               '14:00', '15:00', '15:30', '16:00',
                                                                               '16:30', '17:00', '17:30', '19:30'],
                                                                        '29': ['14:30', '15:00', '15:30', '16:00',
                                                                               '16:30', '17:00', '17:30', '18:00',
                                                                               '18:30', '19:00', '19:30', '20:00',
                                                                               '20:30'],
                                                                        '30': ['08:00', '08:30', '09:00', '09:30',
                                                                               '10:00', '10:30', '11:00', '11:30',
                                                                               '12:00', '12:30', '13:00', '13:30',
                                                                               '14:00', '15:30', '16:00', '16:30',
                                                                               '17:00', '17:30', '18:00', '18:30']}},
                  {'14', '21', '28'}, '/Users/evgeniy/Documents/SMT_clinic/appointments.xlsx')
